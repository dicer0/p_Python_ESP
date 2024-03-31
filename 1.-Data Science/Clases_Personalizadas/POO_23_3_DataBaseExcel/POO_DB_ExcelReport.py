# -*- coding: utf-8 -*-

#pyinstaller nombreArchivo.py --onefile --windowed
#pyinstaller: Es la librería que se debe instalar para convertir un programa de Python a un archivo .exe.
#--onefile: Sirve para que cuando se ejecute el comando de pyinstaller, se cree un solo archivo ejecutable .exe.
#--windowed: Sirve para que cuando se ejecute el ejecutable, no se abra la consola.

#SQLAlchemy: Para instalarlo se debe ejecutar el comando pip install sqlalchemy, y es una biblioteca de Python 
#para SQL ORM (Object-Relacional-Mapping) que proporciona a los desarrolladores un conjunto de herramientas para 
#interactuar con bases de datos utilizando objetos y métodos de Python, en lugar de escribir consultas SQL.
#Entidad: Se refiere a una tabla que almacena datos sobre un tipo de objeto o elemento del mundo real. Cada fila 
#en la tabla representa una instancia individual de esa entidad, y cada columna representa un atributo o 
#característica de esa entidad.
#Atributo: Son las columnas de una tabla que representan las características o propiedades de la entidad que está 
#siendo modelada, todas ellas tienen un nombre y tipo de dato asociado.
#Registro: También conocido como "fila" o "tupla", representa una instancia individual de una entidad en la tabla. 
#Cada registro contiene los valores de los atributos correspondientes a esa instancia específica de la entidad.

#PyODBC: La librería pyodbc permite trabajar con bases de datos utilizando el estándar ODBC (Open DataBase 
#Connectivity), el cual es una API que permite a las aplicaciones conectarse a una amplia variedad de bases de 
#datos que tienen controladores ODBC disponibles, como Microsoft SQL Server, MySQL, PostgreSQL, Oracle, etc.
#Permitiendo a los desarrolladores enviar consultas SQL directamente a la base de datos, sin necesidad de establecer 
#una conexión ODBC.
#No es que exista una mejor librería de manejo de datos, sino que dependiendo de la base de datos a la que nos 
#queramos conectar, a veces conviene utilizar una biblioteca en vez de otra.
# - SQLAlchemy: Conviene usarse cuando se utilizan bases de datos como MySQL Workbench, PostgreSQL y SQLite.
# - PyODBC: Conviene usarse cuando se utilizan bases de datos como Microsoft SQL Server.
import pyodbc
#pandas: Librería que proporciona estructuras de datos y herramientas de manipulación y análisis de datos. 
import pandas
#openpyxl: Esta librería permite leer, escribir y modificar archivos Excel, creando o eliminando sheets, agregando 
#fórmulas de cálculo, dando formato de color o tipo de letra a sus celdas, trabajando con gráficos incrustados, 
#imágenes, etc.   
import openpyxl
#pandas: Librería que proporciona datos adicionales acerca de los errores detectados por una estructura try-except
import traceback

#DatabaseExcelHandler: La clase representa la acción de conectarse a la base de datos a través de una URL, 
#procesar los datos y finalmente crear un reporte en Excel, para ello al constructor se le debe pasar dos 
#parámetros, la URL de conexión a la base de datos y la dirección donde quiero que se guarden mis reportes.
class DatabaseExcelHandler:
    #def __init__(self): Es el constructor o inicializador de la clase, este se llama automáticamente cuando se 
    #crea un objeto que instancíe la clase y en él se declaran los atributos que se reutilizarán en los demás 
    #métodos. En Python, el primer parámetro de cualquier método constructor debe ser self, los demás pueden 
    #servir para cualquier cosa, pero si se declaran en el constructor, estos a fuerza deben tener un valor.
    #self: Se refiere al objeto futuro que se cree a partir de esta clase, es similar al concepto de this en 
    #otros lenguajes de programación.
    def __init__(self, db_connection_string):
        #De esta manera se asigna al atributo self.db_connection_string el valor de la URL que recibe el 
        #constructor de la clase como parámetro.
        self.db_connection_string = db_connection_string    #Atributo: URL de conexión para la base de datos.
        self.connected = False  #Atributo connected: Bandera que indica si la base de datos se pudo conectar.
    
    #Los métodos privados se indican mediante dos guiones bajos al inicio del nombre de la función y esto lo 
    #que implica es que este no se pueda llamar desde afuera de la clase, solo desde otro método internamente.
    def __connect_to_database(self):
        #CONFIGURAR LA CONEXIÓN A DISTINTOS TIPOS DE BASES DE DATOS: Para ello de igual manera se debe realizar 
        #la instalación de diferentes librerías.
        #MANEJO DE EXCEPCIONES: Es una parte de código que se conforma de 2 o 3 partes, try, except y finally: 
        # - Primero se ejecuta el código que haya dentro del try, y si es que llegara a ocurrir una excepción 
        #   durante su ejecución, el programa brinca al código del except.
        # - En la parte de código donde se encuentra la palabra reservada except, se ejecuta cierta acción cuando 
        #   ocurra el error esperado.
        # - Por último, cuando no ocurra una excepción durante la ejecución del gestor de excepciones, se 
        #   ejecutará el código que esté incluido dentro del finally después de haber terminado de ejecutar lo que 
        #   haya en el try, pero si ocurre una excepción, la ejecución terminará cuando se llegue al except.
        #Se utiliza esta arquitectura de código cuando se quiera efectuar una acción donde se espera que pueda 
        #ocurrir un error durante su ejecución.
        #1.- MySQL: create_engine('mysql+pymysql://username:password@hostname:port/database_name')
        #instalation: pip install mysqlclient
        #instalation: pip install pymysql
        try:
            #pyodbc.connect(): Método para establecer la conexión con la base de datos usando PyODBC.
            self.connection1 = pyodbc.connect(self.db_connection_string)
            #pyodbc.connect().cursor(): Una vez que se ha establecido la conexión con el método pyodbc.connect(), 
            #se crea un cursor con el método .cursor(), el cual es un objeto que se utiliza para ejecutar consultas 
            #SQL en la base de datos y recuperar sus resultados, como la inserción, actualización o eliminación de 
            #datos, así como la recuperación de datos de las tablas.
            self.cursor = self.connection1.cursor()
            print("1.- MySQL Connection successful!!!")
            #El atributo connected se cambia manualmente a True cuando se ha realizado la conexión exitosamente.
            self.connected = True
        except Exception as error:
            print("Error occurred while opening the MySQL database:\n" + str(error) + "\n") 

    #Los métodos que no tienen dos guiones bajos al inicio de su nombre son públicos, lo que significa que se 
    #pueden llamar y utilizar desde fuera de la clase. 
    def process_data_and_save_to_excel(self, pathExcel):
        #Ejecución del método privado __connect_to_database() para intentar establecer la conexión a la database.
        self.__connect_to_database()
        #Si no se pudo realizar la conexión con la base de datos, osea cuando la bandera se quede en su valor 
        #predeterminado de False, se devuelve un String que lo indique.
        if self.connected == False:
            return "No se pudo realizar la conexión con la base de datos."
        #Si la variable self.connected cambió su valor a True, se ejecutará un gestor de excepciones para lanzar 
        #un query que lea y procese los datos.
        try:
            #OBTENCIÓN DE DATOS DE LA BASE DE DATOS: Ya que estemos seguros que la conexión a la base de datos 
            #se ha realizado de forma exitosa, podremos utilizar comandos SQL para filtrar y obtener cierta 
            #información, esto se realiza a través de la variable que haya utilizado el método .connect() y el 
            #método .cursor() de la librería pyodbc.
            SQL_Query_string =  """SELECT 	  * 
                                    FROM 	    posts
                                    ORDER BY  titulo DESC;"""
            #pyodbc.connect().cursor().execute(): Ya que se haya realizado la conexión con la base de datos a través 
            #de un objeto cursor, se puede realizar una consulta a la base de datos con SQL.
            self.cursor.execute(SQL_Query_string)
            #pyodbc.connect().cursor().execute().fetchall(): Después de ejecutar la consulta, se llama a este método 
            #en el cursor para recuperar todos los resultados de la consulta, el cual devuelve una lista de tuplas 
            #que contiene todas las filas de resultados de la consulta. Cada fila de la lista es una tupla que 
            #incluye los valores de las columnas de esa fila. Pero cabe mencionar que se deberá realizar una 
            #conversión futura de tupla a tupla, que no cambia nada de los datos, pero sirve para que el programa 
            #identifique que está lidiando con una lista de tuplas, sino cree que es una lista de 1 elemento.
            resultProxy = self.cursor.fetchall()
            print("Tipo de Dato ResultProxy: ", type(resultProxy))
            #TIPOS DE ESTRUCTURAS DE DATOS EN PYTHON: La gran diferencia que estos pueden tener es que algunos 
            #tienen cierto órden (índice y valor) y otros no, además de que algunos son editables o mutables, donde 
            #se les puede agregar, eliminar, o modificar elementos y otros son inmutables, donde sus datos no se 
            #pueden cambiar.
            # - Listas (list): Una lista es una colección ordenada y mutable (editable) de elementos. Se definen 
            #   utilizando corchetes [].
            #       Ejemplo: mi_lista = [1, 2, "hola", True].
            # - Tuplas (tuple): Una tupla es una colección ordenada e inmutable de elementos. Se definen utilizando 
            #   paréntesis ().
            #       Ejemplo: mi_tupla = (1, 2, "hola", True).
            # - Diccionarios (dict): Un diccionario es una colección desordenada y mutable de pares clave-valor. Se 
            #   definen utilizando llaves {} y separando cada par clave-valor por dos puntos :.
            #       Ejemplo: mi_diccionario = {"nombre": "Juan", "edad": 30, "ciudad": "Madrid"}.
            # - Conjuntos (set): Un conjunto es una colección desordenada y mutable de elementos únicos. No permite 
            #   elementos duplicados y no tiene un orden definido. Se definen utilizando llaves {} o utilizando la 
            #   función set().
            #       Ejemplo: mi_conjunto = {1, 2, 3, 4, 5}.
            #ES MUY IMPORTANTE MENCIONAR QUE: Cuando obtenemos datos de una database utilizando PyODBC u otra 
            #librería similar como SQLAlchemy, los datos generalmente se obtienen como una lista de tuplas, donde 
            #cada tupla representa una fila de la tabla. Pero cuando queramos convertir esto a un DataFrame para 
            #manejar mejor los datos al introducirlos en un Excel, debemos realizar de forma explícita la conversión 
            #de las tuplas internas a tuplas, ya que sino el programa podría pensar que es una lista de 1 sola 
            #columna en vez de detectar que se encuentra con una lista de varias columnas, para ello se puede usar 
            #un bucle for de una sola línea, la cual se conforma de la siguiente sintaxis:
            #   resultadoBucleFor = [acción_Realizada      for     variable_Interna_i    in     estructura_De_Datos]
            cursorRows = [tuple(row) for row in resultProxy]            #Conversión explícita de las filas a tuplas.
            #pyodbc.connect().cursor().description: El atributo description perteneciente al objeto cursor de la 
            #librería PyODBC sirve para proporcionar información sobre las columnas de un Query en una tupla de 
            #tuplas, las cuales indican lo siguiente en ese específico órden:
            # - Nombre de la columna: El nombre de la columna tal como está en la base de datos.
            # - Tipo de datos de la columna: El tipo de datos de la columna según la base de datos.
            # - Tamaño de la columna (en bytes): El tamaño máximo de la columna en bytes.
            # - Bytes reservados para la columna: El número de bytes reservados para la columna puede ser None.
            # - Dígitos de precisión: El número de dígitos de precisión para tipos numéricos puede ser None.
            # - Dígitos decimales: El número de dígitos decimales para tipos numéricos puede ser None.
            # - Indicador de si la columna puede ser NULL: Si se puede contener valores NULL (True) o no (False).
            cursorCols = [col[0] for col in self.cursor.description]    #Obtención del título de las columnas.
            #pandas.DataFrame: La clase DataFrame de la librería pandas representa una estructura de datos 
            #matricial en forma de tablas que pueden contener datos de diferentes tipos y se pueden manipular de 
            #manera eficiente para realizar diversas operaciones de análisis de datos, su constructor recibe los 
            #siguientes parámetros:
            # - data: Este es el parámetro principal que especifica los datos que se utilizarán para crear el 
            #   DataFrame. Puede ser un diccionario, una lista de listas, lista de tuplas, un numpyArray, otra 
            #   estructura de datos de Pandas (como otro DataFrame o una Serie), etc.
            # - index (opcional): Este parámetro especifica las etiquetas de índice para las filas del DataFrame. 
            #   Puede ser una lista, una matriz, una Serie, etc. Si no se especifica, se utilizarán índices 
            #   enteros.
            # - columns (opcional): Este parámetro opcional especifica los nombres de las columnas del DataFrame. 
            #   Si no se especifica, se utilizarán nombres de columnas predeterminados (0, 1, 2, ...).
            #       - .keys(): Como los objetos ResultProxy se manejan como diccionarios, se puede obtener el 
            #         nombre de sus etiquetas (keys) para de esa manera asignar los nombres de las columnas de un 
            #         DataFrame, pero si se quisiera obtener sus valores, no existe un método .values(), se debe 
            #         acceder de otra forma.
            # - dtype (opcional): Este parámetro especifica el tipo de datos para cada columna del DataFrame.
            dataFramePandas = pandas.DataFrame(data = cursorRows, columns = cursorCols)
            print(dataFramePandas, "\n")
            #pandas.to_datetime(): Este método se utiliza para convertir un objeto iterable como una lista, tupla, 
            #DataFrame, Serie de Pandas, etc. que contiene fechas o marcas de tiempo en un objeto de tipo DateTime.
            #pandas.to_datetime().dt: El atributo .dt proporciona acceso a una serie de métodos para obtener o 
            #manipular las fechas/horas de su objeto DateTime. Algunas de las cosas que se pueden hacer con el 
            #atributo .dt son:
            # - Fechas/horas individuales: Se puede acceder al año, mes, día, hora, minuto, segundo, etc.
            #       - Obtener el año:               pandas.DataFrame['key_ObjetoDateTime'].dt.year 
            #       - Obtener el mes:               pandas.DataFrame['key_ObjetoDateTime'].dt.month
            #       - Obtener el día:               pandas.DataFrame['key_ObjetoDateTime'].dt.day
            #       - Obtener la hora:              pandas.DataFrame['key_ObjetoDateTime'].dt.hour
            #       - Obtener el día de la semana:  pandas.DataFrame['key_ObjetoDateTime'].dt.dayofweek
            # - Convertir el formato de fecha/hora: Especifica el formato de fecha y hora que se utilizará para 
            #   mostrar el objeto DateTime, para ello se debe proporcionar un código que indique algún formato 
            #   compatible con la sintaxis de strftime de Python:
            #       - %Y: Año con cuatro dígitos (ejemplo: 2022).
            #       - %y: Año con dos dígitos (ejemplo: 22).
            #       - %m: Mes como número decimal (01-12).
            #       - %d: Día del mes como número decimal (01-31).
            #       - %H: Hora (00-23).
            #       - %I: Hora (01-12).
            #       - %p: AM o PM.
            #       - %M: Minuto (00-59).
            #       - %S: Segundo (00-59).
            #       - %f: Microsegundos (000000-999999).
            #       - %j: Día del año (001-366).
            #       - %U: Número de semana del año, comenzando por el domingo (00-53).
            #       - %W: Número de semana del año, comenzando por el lunes (00-53).
            #       - %a: Nombre corto del día de la semana (Sun, Mon, etc.).
            #       - %A: Nombre completo del día de la semana (Sunday, Monday, etc.).
            #       - %b: Nombre corto del mes (Jan, Feb, etc.).
            #       - %B: Nombre completo del mes (January, February, etc.).
            #       - %c: Representación de la fecha y hora local.
            #       - %x: Representación de la fecha local.
            #       - %X: Representación de la hora local.
            # - Calcular diferencias de tiempo: Se pueden calcular diferencias de tiempo entre fechas y horas.
            #       - pandas.DataFrame['DateTime_fecha_fin'] - pandas.DataFrame['DateTime_fecha_inicio']
            # - Agregar o restar intervalos de tiempo: Se puede sumar o restar tiempo a las fechas y horas.
            #       - pandas.DataFrame['DateTime_fecha'] + pandas.Timedelta(days = 1)
            dataFramePandas['fecha_publicacion'] = pandas.to_datetime(dataFramePandas['fecha_publicacion']).dt.strftime('%d-%m-%Y')

            #En este caso lo que se hará es extraer datos de la base de datos, los cuales serán comparados con 
            #algunos valores de la siguiente lista de diccionarios y si algunos de ellos son iguales, se tomará 
            #algunos datos de la database (DB), se extraerán algunos otros de la lista de diccionarios y se 
            #agregarán unos nuevos para crear una nueva estructura de datos, que pueda ser agregada a un reporte 
            #y posteriormente mostrada a su vez en una GUI de PyQt5.
            compareDicc = [{
                "tituloStatic": "Grupo de Datos 1",     #Datos que así se pasan al diccionario final.
                "datoStatic": "Dato grupo 1",         
                "estatusFilter": "activo",              #Datos de filtrado.
                "userIdFilter": 1,
                "categoryIdFilter": 2
            },
            {
                "tituloStatic": "Grupo de Datos 2",     #Datos que así se pasan al diccionario final.
                "datoStatic": "Dato grupo 2",         
                "estatusFilter": "inactivo",            #Datos de filtrado.
                "userIdFilter": 2,
                "categoryIdFilter": 3
            }]            

            #CREAR UN DICCIONARIO QUE COMBINE DATOS DE LA DATABASE CON OTROS A TRAVÉS DE UN DICCIONARIO DE FILTRADO:
            finalData = []  #Diccionario que almacenará los datos que trae PyODBC del DataBase.
            #bucle for each: Es un bucle que recorre la lista de diccionarios compareDicc, por lo que la variable 
            #indDicc lleva contenidos todos los diccionarios en cada iteración, uno por uno.
            for indDicc in compareDicc:
                #Indexación booleana: Es una técnica que se realiza con estructuras de datos de la librería pandas 
                #para filtrar las filas de un DataFrame basándose en el uso de operadores lógicos simples como 
                #& (and), | (or) y ~ (not). La operación solo devolverá las filas del DataFrame que cumplan con 
                #las condiciones descritas dentro de su corchete (osea que devuelvan un True):
                #resultadoFiltrado = pandas.DataFrame()[~(operacionLogica_1 & operacionLogica_2 | operacionLogica_n)]
                #En este punto con el contenido de las filas del DataFrame y la variable indDicc que recorre la 
                #lista de diccionarios es donde se pueden realizar las comparaciones para filtrar los datos.
                filtered = dataFramePandas[(dataFramePandas['estatus'] == indDicc['estatusFilter']) &
                                        (dataFramePandas['usuarios_id'] == indDicc['userIdFilter']) &
                                        (dataFramePandas['categorias_id'] == indDicc['categoryIdFilter'])]
                #pandas.DataFrame().empty: El atributo .empty devuelve un valor True cuando su DataFrame está vacío 
                #y False cuando el DataFrame si cuenta con filas y columnas.
                #El operador not simplemente es una negación (~), osea que cuando algo sea True, lo volverá False 
                #y viceversa, esto se utiliza porque los if se ejecutan cuando su condición vale True.
                if (not(filtered.empty)):
                    #pandas.DataFrame().iterrows(): El método iterrows() se debe aplicar a algun objeto de la clase 
                    #DataFrame y siempre se encontrará como parámetro de un bucle for, ya que este recorre todos los 
                    #datos de su DataFrame, devolviendo una tupla que indica el índice de cada fila y su contenido.
                    for (index, row) in (filtered.iterrows()):
                        #Una vez obtenido el índice y contenido de los elementos del DataFrame, se pueden agregar 
                        #un valor u otro a la lista final dependiendo de si se cumple o no una condición, para ello 
                        #se utilizan condicionales de una sola línea que llevan la siguiente sintaxis:
                        #variable =   valor_si_verdadero        if      condicion       else        valor_si_falso
                        #Si se quiere utilizar una estructura else-if, lo que se hace es agregar un paréntesis y 
                        #varias condiciones de una línea dentro.
                        #variable =   (
                        #               valor_si_verdadero1     if      condicion1      else        valor_si_falso1
                        #               valor_si_verdadero2     if      condicion2      else        valor_si_falso2
                        #               ...
                        #               valor_si_verdadero_n    if      condicion_n     else        valor_si_falso_n
                        #               valor_por_defecto
                        #             )
                        standardContent = """Phasellus laoreet eros nec vestibulum varius. Nunc id efficitur lacus, non imperdiet quam. Aliquam porta, tellus at porta semper, felis velit congue mauris, eu pharetra felis sem vitae tortor. Curabitur bibendum vehicula dolor, nec accumsan tortor ultrices ac. Vivamus nec tristique orci. Nullam fringilla eros magna, vitae imperdiet nisl mattis et. Ut quis malesuada felis. Proin at dictum eros, eget sodales libero. Sed egestas tristique nisi et tempor. Ut cursus sapien eu pellentesque posuere. Etiam eleifend varius cursus.\n\nNullam viverra quam porta orci efficitur imperdiet. Quisque magna erat, dignissim nec velit sit amet, hendrerit mollis mauris. Mauris sapien magna, consectetur et vulputate a, iaculis eget nisi. Nunc est diam, aliquam quis turpis ac, porta mattis neque. Quisque consequat dolor sit amet velit commodo sagittis. Donec commodo pulvinar odio, ut gravida velit pellentesque vitae. Class aptent taciti sociosqu ad litora torquent per conubia nostra, per inceptos himenaeos.\n\nMorbi vulputate ante quis elit pretium, ut blandit felis aliquet. Aenean a massa a leo tristique malesuada. Curabitur posuere, elit sed consectetur blandit, massa mauris tristique ante, in faucibus elit justo quis nisi. Ut viverra est et arcu egestas fringilla. Mauris condimentum, lorem id viverra placerat, libero lacus ultricies est, id volutpat metus sapien non justo. Nulla facilisis, sapien ut vehicula tristique, mauris lectus porta massa, sit amet malesuada dolor justo id lectus. Suspendisse sit amet tempor ligula. Nam sit amet nisl non magna lacinia finibus eget nec augue. Aliquam ornare cursus dapibus. Lorem ipsum dolor sit amet, consectetur adipiscing elit.\n\nDonec ornare sem eget massa pharetra rhoncus. Donec tempor sapien at posuere porttitor. Morbi sodales efficitur felis eu scelerisque. Quisque ultrices nunc ut dignissim vehicula. Donec id imperdiet orci, sed porttitor turpis. Etiam volutpat elit sed justo lobortis, tincidunt imperdiet velit pretium. Ut convallis elit sapien, ac egestas ipsum finibus a. Morbi sed odio et dui tincidunt rhoncus tempor id turpis.\n\nProin fringilla consequat imperdiet. Ut accumsan velit ac augue sollicitudin porta. Phasellus finibus porttitor felis, a feugiat purus tempus vel. Etiam vitae vehicula ex. Praesent ut tellus tellus. Fusce felis nunc, congue ac leo in, elementum vulputate nisi. Duis diam nulla, consequat ac mauris quis, viverra gravida urna."""
                        #En este caso la condición evalúa si en la columna contenido de la base de datos filtrada 
                        #contiene específicamente el valor de "Phasellus laoreet eros nec vestibulum varius..." y si 
                        #es así, coloca un valor que se llama "standard", sino coloca "not conventional".
                        contentStatus = "standard" if (standardContent in row["contenido"]) else "not conventional"
                        #Dentro del bucle, la variable indDicc representa cada diccionario de la lista de 
                        #diccionarios que realiza el filtrado y la variable row representa cada fila de la base 
                        #de datos filtrada.
                        finalData.append({
                            "Titulo Static": indDicc["tituloStatic"],
                            "Content Status": contentStatus,
                            "Dato Static": indDicc["datoStatic"],
                            "Titulo": row["titulo"],
                            "Fecha de Publicacion": row["fecha_publicacion"]
                        })
                #Dentro del else se consideran los datos que no cumplan con las condiciones del filtro, para que 
                #estos no sean despreciados, sino clasificados.
                else:
                    #Una vez teniendo almacenados todos los datos de la base de datos que cumplan las condiciones 
                    #del filtro, se añaden y organizan los datos del DataFrame final que queramos recopilar dentro 
                    #de la lista vacía finalData que no lo hayan hecho.
                    finalData.append({
                        "Titulo Static": "Not categorized",
                        "Content Status": "Not categorized",
                        "Dato Static": "Not categorized",
                        "Titulo": row["titulo"],
                        "Fecha de Publicacion": row["fecha_publicacion"]
                    })
            #Cuando se crea un DataFrame a partir de un diccionario, no es necesario indicar explícitamente las 
            #columnas en su constructor, se pasa directamente a su parámetro data.
            finalDataFrame = pandas.DataFrame(data = finalData)

            #AÑADIR DATOS ESTÁTICOS A UN REPORTE DONDE SE RELLENAN DE FORMA DINÁMICA ALGUNAS TABLAS:
            staticDataAbove_1 = [
                ['Title A1.1', 'Title A1.2', 'Title A1.3'],
                ['Static Row 1', '.', '.'],
                ['Static Row 2', '.', '.'],
                ['Static Row 3', '.', '.'],
                ['Static Row 4', '.', '.'],
                ['Static Row 5', '.', '.'],
                ['Static Row 6', '.', '.']
            ]
            staticDataAbove_2 = [
                ['Title A2', '.', '.', '.', '.', '.', '.'],
                ['Subtitle A2.1', '.', '.', '.', '.', '.', '.'],
                ['Static Row 1', '.', '.', '.', '.', '.', '.'],
                ['.', '.', '.', '.', '.', '.', '.'],
                ['Subtitle A2.2', '.', '.', '.', '.', '.', '.'],
                ['Static Row 2', '.', '.', '.', '.', '.', '.'],
                ['.', '.', '.', '.', '.', '.', '.'],
                ['Subtitle A2.3', '.', '.', '.', '.', '.', '.'],
                ['Static Row 3', '.', '.', '.', '.', '.', '.'],
                ['.', '.', '.', '.', '.', '.', '.'],
                ['Static Text: Lorem ipsum dolor sit amet, consectetur adipiscing elit. Quisque non laoreet mauris. Pellentesque habitant morbi tristique senectus et netus et malesuada fames ac turpis egestas. Curabitur vulputate bibendum nibh elementum pulvinar. Integer a leo in orci ultricies fermentum. Ut vitae velit et sapien congue accumsan sed tincidunt dui. Ut elementum imperdiet nunc, non hendrerit enim ultrices at. Sed rhoncus vehicula.', '.', '.', '.', '.', '.', '.']
            ]
            staticDataBelow_1 = [
                ['Title B1', '.', '.', '.', '.', '.', '.'],
                ['Title B1.1', 'Title B1.2', 'Title B1.3', 'Title B1.4', 'Title B1.5', 'Title B1.6', 'Title B1.7'],
                ['Subtitle 1', '.', '.', '.', '.', '.', '.'],
                ['Static Row 1', '.', '.', '.', '.', '.', '.'],
                ['Static Row 2', '.', '.', '.', '.', '.', '.'],
                ['Static Row 3', '.', '.', '.', '.', '.', '.'],
                ['Static Row 4', '.', '.', '.', '.', '.', '.'],
                ['Static Row 5', '.', '.', '.', '.', '.', '.'],
                ['Static Row 6', '.', '.', '.', '.', '.', '.'],
                ['Static Row 7', '.', '.', '.', '.', '.', '.'],
                ['Subtitle 2', '.', '.', '.', '.', '.', '.'],
                ['Static Row 1', '.', '.', '.', '.', '.', '.']
            ]
            #Para añadir las distintas tablas tanto dinámicas como estáticas se deberá extraer el tamaño del 
            #DataFrame, esto se hará a través de su atributo pandas.DataFrame().shape, el cual devuelve una tupla 
            #que indica su número de filas y columnas: (filas, columnas) = pandas.DataFrame().shape
            (filasDataFrame, columnasDataFrame) = finalDataFrame.shape
            #len(listaDeListas): Para extraer el número de filas de las agrupaciones de datos estáticas se utiliza 
            #el método len(). 
            staticDataAbove_1_Rows = len(staticDataAbove_1)
            staticDataAbove_2_Rows = len(staticDataAbove_2)
            staticDataBelow_1_Rows = len(staticDataBelow_1)
            #len(listaDeListas[posicion]): Para obtener el número de columnas de las agrupaciones de datos estáticas 
            #se usa el método len(), pero aplicado a alguna de las listas internas, para ello se indica cualquier 
            #posición, ya que todas las filas deberían tener el mismo número de columnas (osea de datos).
            staticDataAbove_1_Cols = len(staticDataAbove_1[0])
            staticDataAbove_2_Cols = len(staticDataAbove_2[0])
            staticDataBelow_1_Cols = len(staticDataBelow_1[0])

            #GUARDAR UN DATAFRAME EN UN EXCEL, INDICANDO SU FORMATO ESTÁTICO, NO UNO QUE DEPENDA DE LOS DATOS:
            #openpyxl.Workbook(): El objeto Workbook() de la librería openpyxl sirve para crear el book (archivo de 
            #excel) y sheet (página dentro del book) del Excel para darle formato al archivo. 
            workbook = openpyxl.Workbook()
            #openpyxl.Workbook().active: El atributo active se utiliza para acceder a la hoja de cálculo (sheet) 
            #activa en un libro de trabajo (workbook) creado con openpyxl, ya que como los workbooks contienen una o 
            #más hojas de cálculo, la hoja activa es la que está seleccionada por defecto cuando se abre el libro de 
            #trabajo del Excel.
            worksheet = workbook.active
            #openpyxl.Workbook().active.title = "nombreSheet": Atributo para indicar el título de la página activa 
            #del archivo de Excel (workbook).
            worksheet.title = "Sheet1"

            #RELLENAR EL WORKBOOK DEL EXCEL CON VALORES DINÁMICOS EXTRAÍDOS DE LA DATABASE Y ESTÁTICOS:
            #STATIC DATA ABOVE 1:
            #enumerate(): Es un método que devuelve tanto el índice como el valor de los elementos de una lista, 
            #tupla, diccionario, DataFrame, etc. en forma de una tupla de dos valores (indice, valor).
            #Cuando se quiere acceder a los valores de una lista de listas (tabla), primero se recorren las filas.
            for row_index, row_data in enumerate(staticDataAbove_1):    #Bucle que recorre las filas de la tabla.
                #Y luego se recorren los valores de las columnas, accediendo a las filas de forma individual.
                for col_index, cell_data in enumerate(row_data):        #Bucle que recorre las columnas de la tabla.
                    #openpyxl.Workbook().active.cell(): El método cell() se utiliza para acceder a una celda 
                    #específica en la hoja de cálculo activa de un archivo de Excel, para ello se debe indicar sus 
                    #coordenadas de fila (y_vertical), columna (x_horizontal) y el valor que se le quiere colocar.
                    #Es muy importante mencionar que cuando se utiliza el método enumerate() para extraer los 
                    #índices de las columnas y filas, se les debe sumar un 1 por default, ya que el método .cell()
                    #empieza a contar desde 1 sus coordenadas, no desde 0.
                    row_coordenate = row_index + 1
                    col_coordenate = col_index + 1
                    worksheet.cell(row = row_coordenate, column = col_coordenate, value = cell_data)
            #STATIC DATA ABOVE 2:
            for row_index, row_data in enumerate(staticDataAbove_2):    #Bucle que recorre las filas de la tabla.
                for col_index, cell_data in enumerate(row_data):        #Bucle que recorre las columnas de la tabla.
                    #openpyxl.Workbook().active.cell(): El método cell funciona para colocar datos de forma individual 
                    #en cada celda de la hoja activa de un archivo de Excel.
                    #Se suma +1 por default a las coordenadas porque enumerate() cuenta desde 0 y cell() desde 1.
                    row_coordenate = row_index + 1
                    col_coordenate = col_index + 1
                    worksheet.cell(row = row_coordenate, column = col_coordenate, value = cell_data)
            
            finalDataFrame.to_excel(excel_writer = objetoExcel, index = False, index_label = None, sheet_name = 'Sheet1', startrow = staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1, header = True)
            pandas.DataFrame(staticDataBelow_1).to_excel(excel_writer = objetoExcel, index = False, startrow = staticDataAbove_1_Rows + staticDataAbove_2_Rows + filasDataFrame + 2 + 1 + 1, header = False)

            #Una vez que se haya accedido al book y sheet deseado y guardado ambos en variables, se debe 
            #indicar los formatos estáticos de sus celdas, guardando todos en variables y asignándolos al 
            #book a través del método .add_format().
            #pandas.ExcelWriter().book.add_format({}): Este método sirve para guardar en una variable un 
            #formato de celda, el cual se indica en un diccionario porque se pueden encapsular varias 
            #propiedades:
            # - bg_color: Define el color de fondo de la celda.
            # - font_color: Define el color de la fuente de la celda.
            # - font_name: Define el nombre de la fuente de la celda.
            # - font_size: Define el tamaño de la fuente de la celda.
            # - bold: Define si el texto de la celda está en negrita (True) o no (False).
            # - font_italic: Define si el texto de la celda está en cursiva (True) o no (False).
            # - font_underline: Define si el texto de la celda está subrayado (True) o no (False).
            # - align: Define la alineación del texto dentro de la celda ('left', 'center', 'right', 
            #   'justify', etc.).
            # - valign: Define la alineación vertical del texto dentro de la celda ('top', 'vcenter', 
            #   'bottom', 'vjustify', etc.).
            # - num_format: Define el formato numérico de la celda (por ejemplo, '0.00' para dos decimales).
            # - border: Define los bordes de la celda (puedes especificar si quieres bordes en la parte 
            #   superior, inferior, izquierda, derecha, etc.).
            # - text_wrap: Define si el texto debe envolverse dentro de la celda (True) o no (False).
            #FORMATOS DE COLOR DE LA TABLA ESTÁTICA SUPERIOR 1:
            blueRowDataAbove1_format = workbook.add_format({'bg_color': '#4f81bd'})     #Fila 1 azul.
            blueColDataAbove1_format = workbook.add_format({'bg_color': '#0070c0'})     #Col  2 azul.
            blueTableDataAbove1_format = workbook.add_format({'bg_color': '#d3dfee'})   #Demás celdas azules.
            #FORMATOS DE COLOR DE LA TABLA ESTÁTICA SUPERIOR 2:
            blueRowDataAbove2_format = workbook.add_format({'bg_color': '#4f81bd'})     #Fila 1 azul.
            whiteRowDataAbove2_format = workbook.add_format({'bg_color': 'white'})      #Demás celdas blancas.
            #FORMATOS DE COLOR DE LA TABLA DINÁMICA:
            blueDB_format = workbook.add_format({'bg_color': 'blue'})                   #Fila 1 azul.
            greenDB_format = workbook.add_format({'bg_color': 'green'})                 #Col  1 verde.
            grayDB_format = workbook.add_format({'bg_color': 'gray'})                   #Col  2 gris.
            yellowDB_format = workbook.add_format({'bg_color': 'yellow'})               #Demás celdas amarillas.
            #FORMATOS DE COLOR DE LA TABLA ESTÁTICA INFERIOR 1:
            whiteRowDataBelow1_format = workbook.add_format({'bg_color': 'white'})      #Fila 1 blanca.
            darkBlueRowDataBelow1_format = workbook.add_format({'bg_color': '#4f81bd'}) #Fila 2 azul.
            lightBlueRowDataBelow1_format = workbook.add_format({'bg_color': '#A7BFDE'})#Fila 3 azul claro.
            greenRowDataBelow1_format = workbook.add_format({'bg_color': '#5EC268'})    #Col  1 verde.
            grayRowDataBelow1_format = workbook.add_format({'bg_color': 'gray'})        #Col  2 gris.
            yellowRowDataBelow1_format = workbook.add_format({'bg_color': '#FFF2CC'})   #Demás celdas amarillas.
            #Finalmente se añadirán los formatos previamente guardados en el sheet extraído del objeto 
            #ExcelWriter() a través del método conditional_format().
            #pandas.ExcelWriter().sheets["nombreSheet"].conditional_format(): Método que se utiliza para 
            #aplicar un formato a un rango de celdas en una hoja de cálculo específica (sheet) en un archivo 
            #Excel (book), para ello cabe mencionar que las posiciones de las filas y columnas se empiezan a 
            #contar desde 0 pero su tamaño se considera desde 1.
            # - first_row: Es un número o variable que indica la celda inicial de las filas en la tabla.
            # - first_col: Es un número o variable que indica la celda inicial de las columnas en la tabla.
            # - last_row: Es un número o variable que indica la celda final de las filas en la tabla. Aquí es 
            #   donde se usa la variable filasDataFrame obtenida del atributo pandas.DataFrame().shape.
            # - last_col: Es un número o variable que indica la celda final de las columnas en la tabla. Aquí 
            #   es donde se usa la variable columnasDataFrame obtenida del atributo pandas.DataFrame().shape.
            # - type: Este parámetro indica la condición que se debe cumplir para que se aplique un formato en 
            #   una celda.
            #       - 'cell_value': Aplica el formato si el valor de la celda cumple con cierta condición.
            #       - 'no_blanks': Aplica el formato si la celda no está en blanco.
            #       - 'formula': Aplica el formato si la fórmula en la celda devuelve verdadero.
            #       - 'time_period': Aplica el formato si la fecha en la celda cae dentro de un cierto período 
            #         de tiempo.
            #       - 'text': Aplica el formato basado en el contenido de texto de la celda.
            #       - 'average': Aplica el formato si el valor de la celda es un promedio de un rango.
            #       - 'duplicate': Aplica el formato si el valor de la celda es duplicado de un rango.
            #       - 'unique': Aplica el formato si el valor de la celda es único dentro de un rango.
            #       - '3_color_scale': Aplica un formato de escala de 3 colores basado en los valores de las 
            #         celdas.
            #       - '2_color_scale': Aplica un formato de escala de 2 colores basado en los valores de las 
            #         celdas.
            #       - 'icon_set': Aplica un conjunto de iconos basado en los valores de las celdas.
            # - format: Este parámetro recibe una variable de formato pandas.ExcelWriter().book.add_format({}).
            #FORMATOS DE COLOR DE LA TABLA ESTÁTICA SUPERIOR 1:
            #worksheet.conditional_format(first_row, first_col, last_row, last_col, {'type': 'condition', 'format': formato})
            worksheet.conditional_format(0, 0, 0, (staticDataAbove_1_Cols - 1), {'type': 'no_blanks', 'format': blueRowDataAbove1_format})
            worksheet.conditional_format(1, 0, staticDataAbove_1_Rows, 0, {'type': 'no_blanks', 'format': blueTableDataAbove1_format})
            worksheet.conditional_format(1, 1, staticDataAbove_1_Rows, 1, {'type': 'no_blanks', 'format': blueColDataAbove1_format})
            worksheet.conditional_format(1, 2, staticDataAbove_1_Rows, 2, {'type': 'no_blanks', 'format': blueTableDataAbove1_format})

            #FORMATOS DE COLOR DE LA TABLA ESTÁTICA SUPERIOR 2:
            #pandas.ExcelWriter(engine = 'xlsxwriter').sheets["nombreSheet"].merge_range(): El método 
            #.merge_range() solo se puede utilizar cuando se haya elegido el engine xlsxwriter, y lo que hace 
            #es permitirnos fusionar celdas, para ello puede recibir 3 parámetros:
            # - Código de Celdas de Excel: Recibe el código de celdas de excel de donde a dónde se quiere 
            #   fusionar las celdas, las cuales se indican por medio de una letra y un número, las letras 
            #   indican la columna y el número la fila. Ejemplo: A1:G10 
            #       - Cuando se quiera fusionar celdas conviene concatenar el número de la fila con el nombre 
            #         de la celda, ya que de esta manera se puede automatizar el proceso cuando se busca 
            #         fusionar las celdas de más de 1 fila.
            # - data: Este parámetro recibe el texto (dato) que se va a mostrar en la celda fusionada, si no se 
            #   quiere añadir ningún texto, se deja como None.
            # - format (opcional): Se puede crear un formato de letra para que se aplique a esta celda.
            rowPositionStaticDataAbove2 = (staticDataAbove_1_Rows + 1) + 1
            ExcelCellsStaticDataAbove2 = "A" + str(rowPositionStaticDataAbove2) + ":G" + str(rowPositionStaticDataAbove2)    #Position  A9:G9
            worksheet.merge_range(ExcelCellsStaticDataAbove2, data = None)
            #pandas.ExcelWriter().sheets["nombreSheet"].conditional_format(fila_inicial, col_inicial, fila_final, col_final, {type})
            worksheet.conditional_format((staticDataAbove_1_Rows + 1), 0, (staticDataAbove_1_Rows + 1), (staticDataAbove_2_Cols - 1), {'type': 'no_blanks', 'format': blueRowDataAbove2_format})
            for i in range(0, staticDataAbove_2_Rows - 1): #Position  A10:G10 - A19:G19
                rowPositionStaticDataAbove2 += 1
                ExcelCellsStaticDataAbove2 = "A" + str(rowPositionStaticDataAbove2) + ":G" + str(rowPositionStaticDataAbove2)
                worksheet.merge_range(ExcelCellsStaticDataAbove2, data = None)
                worksheet.conditional_format((staticDataAbove_1_Rows + 2), 0, (staticDataAbove_1_Rows + 2), (staticDataAbove_2_Cols - 1), {'type': 'no_blanks', 'format': whiteRowDataAbove2_format})
            
            #FORMATOS DE COLOR DE LA TABLA DINÁMICA:
            #pandas.ExcelWriter().sheets["nombreSheet"].conditional_format(fila_inicial, col_inicial, fila_final, col_final, {type})
            worksheet.conditional_format((staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1), 0, (staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1), (columnasDataFrame - 1), {'type': 'no_blanks', 'format': blueDB_format})
            worksheet.conditional_format(((staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1) + 1), 0, filasDataFrame + ((staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1) + 1), 0, {'type': 'no_blanks', 'format': greenDB_format})
            worksheet.conditional_format(((staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1) + 1), 1, filasDataFrame + ((staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1) + 1), 1, {'type': 'no_blanks', 'format': grayDB_format})
            for col in range(2, columnasDataFrame):
                worksheet.conditional_format(((staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1) + 1), col, filasDataFrame + ((staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1 + 1) + 1), col, {'type': 'no_blanks', 'format': yellowDB_format})
            
            #FORMATOS DE COLOR DE LA TABLA ESTÁTICA INFERIOR 1:
            #pandas.ExcelWriter().sheets["nombreSheet"].conditional_format(fila_inicial, col_inicial, fila_final, col_final, {type})
            rowPositionStaticDataBelow1 = (staticDataAbove_1_Rows + staticDataAbove_2_Rows + filasDataFrame + 1 + 1 + 1 + 1) + 1
            ExcelCell = "A" + str(rowPositionStaticDataBelow1) + ":G" + str(rowPositionStaticDataBelow1)    #Position  A29:G29
            worksheet.merge_range(ExcelCell, data = None)
            worksheet.conditional_format((rowPositionStaticDataBelow1 - 1), 0, (rowPositionStaticDataBelow1 - 1), (staticDataBelow_1_Cols - 1), {'type': 'no_blanks', 'format': whiteRowDataBelow1_format})
            worksheet.conditional_format(rowPositionStaticDataBelow1, 0, rowPositionStaticDataBelow1, (staticDataBelow_1_Cols - 1), {'type': 'no_blanks', 'format': darkBlueRowDataBelow1_format})
            rowPositionStaticDataBelow1 += 2
            ExcelCell = "A" + str(rowPositionStaticDataBelow1) + ":G" + str(rowPositionStaticDataBelow1)    #Position  A31:G31
            worksheet.merge_range(ExcelCell, data = None)
            worksheet.conditional_format((rowPositionStaticDataBelow1 - 1), 0, (rowPositionStaticDataBelow1 - 1), (staticDataBelow_1_Cols - 1), {'type': 'no_blanks', 'format': lightBlueRowDataBelow1_format})
            worksheet.conditional_format(rowPositionStaticDataBelow1, 0, (staticDataBelow_1_Rows + rowPositionStaticDataBelow1), 0, {'type': 'no_blanks', 'format': greenRowDataBelow1_format})
            worksheet.conditional_format(rowPositionStaticDataBelow1, 1, (staticDataBelow_1_Rows + rowPositionStaticDataBelow1), 1, {'type': 'no_blanks', 'format': grayRowDataBelow1_format})
            for col in range(2, staticDataBelow_1_Cols):
                worksheet.conditional_format(rowPositionStaticDataBelow1, col, (staticDataBelow_1_Rows + rowPositionStaticDataBelow1), col, {'type': 'no_blanks', 'format': yellowRowDataBelow1_format})
        
            #AJUSTAR LAS DIMENSIONES DE LAS CELDAS PARA MOSTRAR SU CONTENIDO COMPLETO:  
            max_lengths = [len(str(col)) for col in finalDataFrame.columns]
            for index, row in finalDataFrame.iterrows():
                for i, value in enumerate(row):
                    max_lengths[i] = max(max_lengths[i], len(str(value)))
            #Aplicar los anchos máximos a las columnas del DataFrame
            for i, max_length in enumerate(max_lengths):
                worksheet.set_column(i, i, max_length + 1)  #Agregar un margen de 1 para mejor aspecto
            
            #Devolver el DataFrame procesado.
            return finalDataFrame
        except Exception as error:
            print("1.- Ups an Error ocurred while Opening the MySQL DataBase:\n" + str(error) + "\n")
            print("La línea donde ocurrió el error fue: ", traceback.format_exc())
            return "Error al procesar los datos y guardarlos en un Excel."
        finally:
            if self.connection1:
                self.connection1.close()
                print("MySQL Connection closed.")