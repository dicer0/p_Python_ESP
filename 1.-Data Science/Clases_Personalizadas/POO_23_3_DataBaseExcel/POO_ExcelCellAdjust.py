# -*- coding: utf-8 -*-

#openpyxl: Librería que permite crear, leer, escribir y modificar archivos de Excel (.xlsx).
import openpyxl

#ExcelCellAdjuster: Clase propia que ajusta el ancho de las celdas de un Excel en función de su contenido.
class ExcelCellAdjuster:
    #def __init__(self): Es el constructor o inicializador de la clase, este se llama automáticamente cuando se 
    #crea un objeto que instancíe la clase y en él se declaran los atributos que se reutilizarán en los demás 
    #métodos. En Python, el primer parámetro de cualquier método constructor debe ser self, los demás pueden 
    #servir para cualquier cosa, pero si se declaran en el constructor, estos a fuerza deben tener un valor.
    #self: Se refiere al objeto futuro que se cree a partir de esta clase, es similar al concepto de this en 
    #otros lenguajes de programación.
    def __init__(self, ruta_excel, ancho_maximo, altura_maxima):
        #De esta manera se asignan valores a los atributos que recibe el constructor de la clase como parámetro:
        self.ruta_excel = ruta_excel        #Path del archivo de excel.
        self.ancho_maximo = ancho_maximo    #Ancho máximo de las celdas en el Excel después del ajuste.
        self.altura_maxima = altura_maxima  #Altura máxima de las celdas en el Excel después del ajuste.

    #ajustar_celdas(): Método que recibe todos los parámetros del constructor y trabaja con ellos para 
    #ajustar de forma automática el tamaño de las celdas del Excel, pero al mismo tiempo limitar dicho tamaño.
    def ajustar_celdas(self):
        #openpyxl.load_workbook(): Método para cargar un archivo (libro) de Excel para acceder a sus hojas de 
        #cálculo, leer, modificar celdas, trabajar con gráficos, estilos, rangos de celdas fusionadas, etc.
        wb = openpyxl.load_workbook(self.ruta_excel)
        #openpyxl.load_workbook().active: El atributo active permite el acceso a la hoja de cálculo activa 
        #(principal o actual) dentro del workbook cargado con el método  openpyxl.load_workbook().
        ws = wb.active

        #Cuando se tengan filas fusionadas en el archivo de Excel, vale la pena manejar la primera columna 
        #por separado, ya que si el contenido de alguna de las filas es demasiado largo, el ancho de esta no 
        #se ajustará correctamente.
        max_length_first_col = 0
        for cell in ws[ws.cell(row=1, column=1).column]:
            try:
                if len(str(cell.value)) > max_length_first_col:
                    max_length_first_col = len(str(cell.value))
            except:
                pass
        adjusted_width_first_col = min((max_length_first_col + 2) * 1.2, self.ancho_maximo)
        ws.column_dimensions['A'].width = adjusted_width_first_col

        # Adjust the width of the remaining columns based on the content
        for col in list(ws.columns)[1:]:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, self.ancho_maximo)
            ws.column_dimensions[column].width = adjusted_width

        # Adjust the height of the rows based on the content
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                try:
                    lines = str(cell.value).count('\n') + 1
                    height = lines * 15
                    if height > max_height:
                        max_height = height
                except:
                    pass
            # Set a default value if the height is None
            if max_height == 0:
                max_height = 15  # Default value
            adjusted_height = min(max_height, self.altura_maxima)
            ws.row_dimensions[row[0].row].height = adjusted_height

        # Save the workbook with the adjustments made
        wb.save(self.ruta_excel)

# Usage:
excelFilePath2 = "C:/Users/diego/OneDrive/Documents/The_MechaBible/p_Python_ESP/1.-Data Science/0.-Archivos_Ejercicios_Python/23.-GUI PyQt5 Conexion DataBase/23.-Reporte Analisis de Datos 1.xlsx"
ancho_maximo = 40
altura_maxima = 20
# Create an instance of the ExcelCellAdjuster class
adjuster = ExcelCellAdjuster(excelFilePath2, ancho_maximo, altura_maxima)
# Call the method to adjust the cells
adjuster.ajustar_celdas()