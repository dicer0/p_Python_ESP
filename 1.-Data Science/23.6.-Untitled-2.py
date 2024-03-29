# -*- coding: utf-8 -*-
import pyodbc
import pandas
import traceback
import openpyxl

class DatabaseExcelHandler:
    def __init__(self, db_connection_string):
        self.db_connection_string = db_connection_string
        self.connected = False
    
    def __connect_to_database(self):
        try:
            self.connection1 = pyodbc.connect(self.db_connection_string)
            self.cursor = self.connection1.cursor()
            print("1.- MySQL Connection successful!!!")
            self.connected = True
        except Exception as error:
            print("Error occurred while opening the MySQL database:\n" + str(error) + "\n") 

    def process_data_and_save_to_excel(self, pathExcel):
        self.__connect_to_database()
        if self.connected == False:
            return "No se pudo realizar la conexión con la base de datos."
        try:
            SQL_Query_string =  """SELECT 	  * 
                                    FROM 	    posts
                                    ORDER BY  titulo DESC;"""
            self.cursor.execute(SQL_Query_string)
            resultProxy = self.cursor.fetchall()
            print("Tipo de Dato ResultProxy: ", type(resultProxy))
            cursorRows = [tuple(row) for row in resultProxy]
            cursorCols = [col[0] for col in self.cursor.description]
            dataFramePandas = pandas.DataFrame(data=cursorRows, columns=cursorCols)
            print(dataFramePandas, "\n")
            dataFramePandas['fecha_publicacion'] = pandas.to_datetime(dataFramePandas['fecha_publicacion']).dt.strftime('%d-%m-%Y')
            compareDicc = [{
                "tituloStatic": "Grupo de Datos 1",
                "datoStatic": "Dato grupo 1",         
                "estatusFilter": "activo",
                "userIdFilter": 1,
                "categoryIdFilter": 2
            },
            {
                "tituloStatic": "Grupo de Datos 2",
                "datoStatic": "Dato grupo 2",         
                "estatusFilter": "inactivo",
                "userIdFilter": 2,
                "categoryIdFilter": 3
            }]
            finalData = []
            for indDicc in compareDicc:
                filtered = dataFramePandas[(dataFramePandas['estatus'] == indDicc['estatusFilter']) &
                                        (dataFramePandas['usuarios_id'] == indDicc['userIdFilter']) &
                                        (dataFramePandas['categorias_id'] == indDicc['categoryIdFilter'])]
                if (not(filtered.empty)):
                    for index, row in filtered.iterrows():
                        standardContent = """Phasellus laoreet eros nec vestibulum varius. Nunc id efficitur lacus, non imperdiet quam. Aliquam porta, tellus at porta semper, felis velit congue mauris, eu pharetra felis sem vitae tortor. Curabitur bibendum vehicula dolor, nec accumsan tortor ultrices ac. Vivamus nec tristique orci. Nullam fringilla eros magna, vitae imperdiet nisl mattis et. Ut quis malesuada felis. Proin at dictum eros, eget sodales libero. Sed egestas tristique nisi et tempor. Ut cursus sapien eu pellentesque posuere. Etiam eleifend varius cursus.\n\nNullam viverra quam porta orci efficitur imperdiet. Quisque magna erat, dignissim nec velit sit amet, hendrerit mollis mauris. Mauris sapien magna, consectetur et vulputate a, iaculis eget nisi. Nunc est diam, aliquam quis turpis ac, porta mattis neque. Quisque consequat dolor sit amet velit commodo sagittis. Donec commodo pulvinar odio, ut gravida velit pellentesque vitae. Class aptent taciti sociosqu ad litora torquent per conubia nostra, per inceptos himenaeos.\n\nMorbi vulputate ante quis elit pretium, ut blandit felis aliquet. Aenean a massa a leo tristique malesuada. Curabitur posuere, elit sed consectetur blandit, massa mauris tristique ante, in faucibus elit justo quis nisi. Ut viverra est et arcu egestas fringilla. Mauris condimentum, lorem id viverra placerat, libero lacus ultricies est, id volutpat metus sapien non justo. Nulla facilisis, sapien ut vehicula tristique, mauris lectus porta massa, sit amet malesuada dolor justo id lectus. Suspendisse sit amet tempor ligula. Nam sit amet nisl non magna lacinia finibus eget nec augue. Aliquam ornare cursus dapibus. Lorem ipsum dolor sit amet, consectetur adipiscing elit.\n\nDonec ornare sem eget massa pharetra rhoncus. Donec tempor sapien at posuere porttitor. Morbi sodales efficitur felis eu scelerisque. Quisque ultrices nunc ut dignissim vehicula. Donec id imperdiet orci, sed porttitor turpis. Etiam volutpat elit sed justo lobortis, tincidunt imperdiet velit pretium. Ut convallis elit sapien, ac egestas ipsum finibus a. Morbi sed odio et dui tincidunt rhoncus tempor id turpis.\n\nProin fringilla consequat imperdiet. Ut accumsan velit ac augue sollicitudin porta. Phasellus finibus porttitor felis, a feugiat purus tempus vel. Etiam vitae vehicula ex. Praesent ut tellus tellus. Fusce felis nunc, congue ac leo in, elementum vulputate nisi. Duis diam nulla, consequat ac mauris quis, viverra gravida urna."""
                        contentStatus = "standard" if (standardContent in row["contenido"]) else "not conventional"
                        finalData.append({
                            "Titulo Static": indDicc["tituloStatic"],
                            "Content Status": contentStatus,
                            "Dato Static": indDicc["datoStatic"],
                            "Titulo": row["titulo"],
                            "Fecha de Publicacion": row["fecha_publicacion"]
                        })
                else:
                    finalData.append({
                        "Titulo Static": "Not categorized",
                        "Content Status": "Not categorized",
                        "Dato Static": "Not categorized",
                        "Titulo": row["titulo"],
                        "Fecha de Publicacion": row["fecha_publicacion"]
                    })
            finalDataFrame = pandas.DataFrame(data=finalData)

            staticDataAbove_1 = [
                ['Title A1.1', 'Title A1.2', 'Title A1.3'],
                ['Static Row 1', '', ''],
                ['Static Row 2', '', ''],
                ['Static Row 3', '', ''],
                ['Static Row 4', '', ''],
                ['Static Row 5', '', ''],
                ['Static Row 6', '', '']
            ]
            staticDataAbove_2 = [
                ['Title A2', '', '', '', '', '', ''],
                ['Subtitle A2.1', '', '', '', '', '', ''],
                ['Static Row 1', '', '', '', '', '', ''],
                ['', '', '', '', '', '', ''],
                ['Subtitle A2.2', '', '', '', '', '', ''],
                ['Static Row 2', '', '', '', '', '', ''],
                ['', '', '', '', '', '', ''],
                ['Subtitle A2.3', '', '', '', '', '', ''],
                ['Static Row 3', '', '', '', '', '', ''],
                ['', '', '', '', '', '', ''],
                ['Static Text: Lorem ipsum dolor sit amet, consectetur adipiscing elit. Quisque non laoreet mauris. Pellentesque habitant morbi tristique senectus et netus et malesuada fames ac turpis egestas. Curabitur vulputate bibendum nibh elementum pulvinar. Integer a leo in orci ultricies fermentum. Ut vitae velit et sapien congue accumsan sed tincidunt dui. Ut elementum imperdiet nunc, non hendrerit enim ultrices at. Sed rhoncus vehicula.', '', '', '', '', '', '']
            ]
            staticDataBelow_1 = [
                ['Title B1', '', '', '', '', '', ''],
                ['Title B1.1', 'Title B1.2', 'Title B1.3', 'Title B1.4', 'Title B1.5', 'Title B1.6', 'Title B1.7'],
                ['Subtitle 1', '', '', '', '', '', ''],
                ['Static Row 1', '', '', '', '', '', ''],
                ['Static Row 2', '', '', '', '', '', ''],
                ['Static Row 3', '', '', '', '', '', ''],
                ['Static Row 4', '', '', '', '', '', ''],
                ['Static Row 5', '', '', '', '', '', ''],
                ['Static Row 6', '', '', '', '', '', ''],
                ['Static Row 7', '', '', '', '', '', ''],
                ['Subtitle 2', '', '', '', '', '', ''],
                ['Static Row 1', '', '', '', '', '', '']
            ]

            (filasDataFrame, columnasDataFrame) = finalDataFrame.shape
            staticDataAbove_1_Rows = len(staticDataAbove_1)
            staticDataAbove_2_Rows = len(staticDataAbove_2)
            staticDataBelow_1_Rows = len(staticDataBelow_1)
            staticDataAbove_1_Cols = len(staticDataAbove_1[0])
            staticDataAbove_2_Cols = len(staticDataAbove_2[0])
            staticDataBelow_1_Cols = len(staticDataBelow_1[0])
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Sheet1'

            blueRowDataAbove1_format = openpyxl.styles.PatternFill(start_color="4f81bd", end_color="4f81bd", fill_type="solid")
            blueTableDataAbove1_format = openpyxl.styles.PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
            blueColDataAbove1_format = openpyxl.styles.PatternFill(start_color="dae3f5", end_color="dae3f5", fill_type="solid")
            whiteRowDataAbove2_format = openpyxl.styles.PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
            blueRowDataAbove2_format = openpyxl.styles.PatternFill(start_color="9dc3e6", end_color="9dc3e6", fill_type="solid")
            whiteRowDataBelow1_format = openpyxl.styles.PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
            darkBlueRowDataBelow1_format = openpyxl.styles.PatternFill(start_color="284157", end_color="284157", fill_type="solid")
            lightBlueRowDataBelow1_format = openpyxl.styles.PatternFill(start_color="cfe2f3", end_color="cfe2f3", fill_type="solid")
            greenRowDataBelow1_format = openpyxl.styles.PatternFill(start_color="6fa53e", end_color="6fa53e", fill_type="solid")
            grayRowDataBelow1_format = openpyxl.styles.PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")
            yellowRowDataBelow1_format = openpyxl.styles.PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
            blueDB_format = openpyxl.styles.PatternFill(start_color="a9d18e", end_color="a9d18e", fill_type="solid")
            greenDB_format = openpyxl.styles.PatternFill(start_color="6fa53e", end_color="6fa53e", fill_type="solid")
            grayDB_format = openpyxl.styles.PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")
            yellowDB_format = openpyxl.styles.PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")

            # Writing staticDataAbove_1
            for row_index, row_data in enumerate(staticDataAbove_1):
                for col_index, cell_data in enumerate(row_data):
                    worksheet.cell(row=row_index + 1, column=col_index + 1, value=cell_data)
            
            # Writing staticDataAbove_2
            for row_index, row_data in enumerate(staticDataAbove_2):
                for col_index, cell_data in enumerate(row_data):
                    worksheet.cell(row=row_index + staticDataAbove_1_Rows + 1, column=col_index + 1, value=cell_data)

            # Writing finalDataFrame
            for row_index, row_data in finalDataFrame.iterrows():
                for col_index, cell_data in enumerate(row_data):
                    worksheet.cell(row=row_index + staticDataAbove_1_Rows + staticDataAbove_2_Rows + 1, column=col_index + 1, value=cell_data)

            # Writing staticDataBelow_1
            for row_index, row_data in enumerate(staticDataBelow_1):
                for col_index, cell_data in enumerate(row_data):
                    worksheet.cell(row=row_index + staticDataAbove_1_Rows + staticDataAbove_2_Rows + filasDataFrame + 2, column=col_index + 1, value=cell_data)
            
            # Applying cell background colors
            fill = openpyxl.styles.PatternFill(start_color="4f81bd", end_color="4f81bd", fill_type="solid")
            for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
                for cell in row:
                    cell.fill = fill

            # Applying conditional formatting
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=blueRowDataAbove1_format)
            worksheet.conditional_formatting.add('A1:C' + str(staticDataAbove_1_Rows), rule)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=blueColDataAbove1_format)
            worksheet.conditional_formatting.add('B1:B' + str(staticDataAbove_1_Rows), rule)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=blueTableDataAbove1_format)
            worksheet.conditional_formatting.add('C1:C' + str(staticDataAbove_1_Rows), rule)
            
            rowPositionStaticDataAbove2 = staticDataAbove_1_Rows + 2
            ExcelCellsStaticDataAbove2 = "A" + str(rowPositionStaticDataAbove2) + ":G" + str(rowPositionStaticDataAbove2)
            worksheet.merge_cells(ExcelCellsStaticDataAbove2)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=blueRowDataAbove2_format)
            worksheet.conditional_formatting.add(ExcelCellsStaticDataAbove2, rule)

            for i in range(1, staticDataAbove_2_Rows):
                rowPositionStaticDataAbove2 += 1
                ExcelCellsStaticDataAbove2 = "A" + str(rowPositionStaticDataAbove2) + ":G" + str(rowPositionStaticDataAbove2)
                worksheet.merge_cells(ExcelCellsStaticDataAbove2)
                rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=whiteRowDataAbove2_format)
                worksheet.conditional_formatting.add(ExcelCellsStaticDataAbove2, rule)

            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=blueDB_format)
            worksheet.conditional_formatting.add('A' + str(staticDataAbove_1_Rows + staticDataAbove_2_Rows + 2) + ':' + chr(ord('A') + columnasDataFrame - 1) + str(staticDataAbove_1_Rows + staticDataAbove_2_Rows + 2), rule)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=greenDB_format)
            worksheet.conditional_formatting.add('A' + str(staticDataAbove_1_Rows + staticDataAbove_2_Rows + 3) + ':' + 'A' + str(filasDataFrame + staticDataAbove_1_Rows + staticDataAbove_2_Rows + 3), rule)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=grayDB_format)
            worksheet.conditional_formatting.add('B' + str(staticDataAbove_1_Rows + staticDataAbove_2_Rows + 3) + ':' + 'B' + str(filasDataFrame + staticDataAbove_1_Rows + staticDataAbove_2_Rows + 3), rule)
            for col in range(2, columnasDataFrame):
                rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=yellowDB_format)
                worksheet.conditional_formatting.add(chr(ord('A') + col) + str(staticDataAbove_1_Rows + staticDataAbove_2_Rows + 3) + ':' + chr(ord('A') + col) + str(filasDataFrame + staticDataAbove_1_Rows + staticDataAbove_2_Rows + 3), rule)

            rowPositionStaticDataBelow1 = staticDataAbove_1_Rows + staticDataAbove_2_Rows + filasDataFrame + 4
            ExcelCell = "A" + str(rowPositionStaticDataBelow1) + ":G" + str(rowPositionStaticDataBelow1)
            worksheet.merge_cells(ExcelCell)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=whiteRowDataBelow1_format)
            worksheet.conditional_formatting.add('A' + str(rowPositionStaticDataBelow1 - 1) + ':' + chr(ord('G')) + str(rowPositionStaticDataBelow1 - 1), rule)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=darkBlueRowDataBelow1_format)
            worksheet.conditional_formatting.add('A' + str(rowPositionStaticDataBelow1) + ':' + chr(ord('G')) + str(rowPositionStaticDataBelow1), rule)

            rowPositionStaticDataBelow1 += 2
            ExcelCell = "A" + str(rowPositionStaticDataBelow1) + ":G" + str(rowPositionStaticDataBelow1)
            worksheet.merge_cells(ExcelCell)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=lightBlueRowDataBelow1_format)
            worksheet.conditional_formatting.add('A' + str(rowPositionStaticDataBelow1 - 1) + ':' + chr(ord('G')) + str(rowPositionStaticDataBelow1 - 1), rule)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=greenRowDataBelow1_format)
            worksheet.conditional_formatting.add('A' + str(rowPositionStaticDataBelow1) + ':' + 'A' + str(rowPositionStaticDataBelow1 + staticDataBelow_1_Rows - 1), rule)
            rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=grayRowDataBelow1_format)
            worksheet.conditional_formatting.add('B' + str(rowPositionStaticDataBelow1) + ':' + 'B' + str(rowPositionStaticDataBelow1 + staticDataBelow_1_Rows - 1), rule)
            for col in range(2, staticDataBelow_1_Cols):
                rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual', formula=['" "'], stopIfTrue=True, fill=yellowRowDataBelow1_format)
                worksheet.conditional_formatting.add(chr(ord('A') + col) + str(rowPositionStaticDataBelow1) + ':' + chr(ord('A') + col) + str(rowPositionStaticDataBelow1 + staticDataBelow_1_Rows - 1), rule)

            workbook.save(pathExcel)
            print("Database table data successfully written to Excel file")
            return "Proceso de exportación de datos a Excel finalizado con éxito."
        except Exception as e:
            print("Error occurred while processing data and saving to Excel:\n" + str(e) + "\n")
            print(traceback.format_exc())
            return "Se produjo un error al procesar los datos y guardar en Excel."

if __name__ == "__main__":
    connectionString = 'DRIVER={MySQL ODBC 8.3 Unicode Driver};SERVER=localhost;PORT=3306;DATABASE=1_platziblog_db;USER=root;PASSWORD=PincheTonto!123;'
    db_handler1 = DatabaseExcelHandler(connectionString)
    excelFilePath2 = "C:/Users/diego/OneDrive/Documents/The_MechaBible/p_Python_ESP/1.-Data Science/0.-Archivos_Ejercicios_Python/23.-GUI PyQt5 Conexion DataBase/23.-Reporte Analisis de Datos 2.xlsx"
    db_handler1.process_data_and_save_to_excel(excelFilePath2)