import openpyxl

def ajustar_celdas_excel(ruta_excel, ancho_maximo, altura_maxima):
    excelWorkBook = openpyxl.load_workbook(ruta_excel)
    excelWorkSheet = excelWorkBook.active
    
    numLetrasCol_1 = 0
    cellObject = excelWorkSheet.cell(row = 1, column = 1)
    column1Index = cellObject.column
    for cell in excelWorkSheet[column1Index]:
        if ((cell.value is not None) and (len(cell.value) > numLetrasCol_1)):
            numLetrasCol_1 = len(cell.value)
    anchoMaxCol1 = (numLetrasCol_1 + 2) * 1.4
    adjusted_width_first_col = min(anchoMaxCol1, ancho_maximo)
    excelWorkSheet.column_dimensions['A'].width = adjusted_width_first_col

    columns_list = list(excelWorkSheet.columns)
    for col in columns_list[1:]:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if ((cell.value is not None) and (len(cell.value) > max_length)):
                max_length = len(cell.value)
        adjusted_width = min((max_length + 2) * 1.2, ancho_maximo)
        excelWorkSheet.column_dimensions[column].width = adjusted_width

    for row in excelWorkSheet.iter_rows():
        max_height = 0
        for cell in row:
            lines = str(cell.value).count('\n') + 1
            height = lines * 15
            if (height > max_height):
                max_height = height
        for cell in row:
            if (cell.coordinate in excelWorkSheet.merged_cells):
                for range_ in excelWorkSheet.merged_cells.ranges:
                    if (cell.coordinate in range_):
                        for row_ in range(range_.min_row, range_.max_row + 1):
                            if (excelWorkSheet.row_dimensions[row_].height is not None):
                                max_height = max(max_height, excelWorkSheet.row_dimensions[row_].height)
        if (max_height == 0):
            max_height = 15
        adjusted_height = min(max_height, altura_maxima)
        excelWorkSheet.row_dimensions[row[0].row].height = adjusted_height

    excelWorkBook.save(ruta_excel)

excelFilePath2 = "C:/Users/diego/OneDrive/Documents/The_MechaBible/p_Python_ESP/1.-Data Science/0.-Archivos_Ejercicios_Python/23.-GUI PyQt5 Conexion DataBase/23.-Reporte Analisis de Datos 2.xlsx"
ancho_maximo = 40
altura_maxima = 20
ajustar_celdas_excel(excelFilePath2, ancho_maximo, altura_maxima)