# -*- coding: utf-8 -*-

#openpyxl: Librería que permite crear, leer, escribir y modificar archivos de Excel (.xlsx).
import openpyxl

#ExcelCellAdjuster: Clase propia que ajusta el ancho de las celdas de un Excel en función de su contenido.
class ExcelCellAdjuster:
    #def __init__(self): Es el constructor o inicializador de la clase, este se llama automáticamente cuando se 
    #crea un objeto que instancíe la clase y en él se declaran los atributos que se reutilizarán en los demás 
    #métodos. En Python, el primer parámetro de cualquier método constructor debe ser self, los demás pueden 
    #servir para cualquier cosa, pero si se declaran en el constructor, estos a fuerza deben tener un valor.
    #self: Se refiere al objeto futuro que se cree a partir de esta clase, es similar al concepto de this en 
    #otros lenguajes de programación.
    def __init__(self, ruta_excel, ancho_maximo):
        #De esta manera se asignan valores a los atributos que recibe el constructor de la clase como parámetro:
        self.ruta_excel = ruta_excel        #Path del archivo de excel.
        self.ancho_maximo = ancho_maximo    #Ancho máximo de las celdas en el Excel después del ajuste.

    #ajustar_celdas(): Método que recibe todos los parámetros del constructor y trabaja con ellos para 
    #ajustar de forma automática el tamaño de las celdas del Excel, pero al mismo tiempo limitar dicho tamaño.
    def ajustar_celdas(self):
        #openpyxl.load_workbook(): Método para cargar un archivo (libro) de Excel para acceder a sus hojas de 
        #cálculo, leer, modificar celdas, trabajar con gráficos, estilos, rangos de celdas fusionadas, etc.
        excelWorkBook = openpyxl.load_workbook(self.ruta_excel)
        #openpyxl.load_workbook().active: El atributo active permite el acceso a la hoja de cálculo activa 
        #(principal o actual) dentro del workbook cargado con el método  openpyxl.load_workbook(). La hoja activa 
        #es aquella que estaba abierta cuando el libro de Excel se guardó por última vez o que está actualmente 
        #seleccionada dentro del libro de Excel.
        excelWorkSheet = excelWorkBook.active

        #Cuando se tengan filas fusionadas en el archivo de Excel, vale la pena manejar la primera columna 
        #por separado, ya que si el contenido de alguna de las filas es demasiado largo, el ancho de esta no 
        #se ajustará correctamente.
        #openpyxl.load_workbook().active.cell(row, column): El método cell() se aplica al worksheet activo de un 
        #workbook (archivo de Excel) y devuelve un objeto openpyxl.Cell, que representa una sola celda de la 
        #tabla, para así poder acceder o maniuplar sus propiedades. Este recibe los parámetros row y column, 
        #donde se indica la coordenada de la celda que queremos acceder contando desde 1. 
        cellObject = excelWorkSheet.cell(row = 1, column = 1)
        #openpyxl.Cell: Los atributos del objeto Cell que devuelven propiedades específicas son:
        # - openpyxl.Cell.value:        Atributo que devuelve el valor contenido en la celda en forma de string.
        # - openpyxl.Cell.row:          Devuelve el número de fila de la celda.
        # - openpyxl.Cell.column:       Devuelve el índice de la columna a la que pertenece la celda contando 
        #   desde 1.
        # - openpyxl.Cell.column_letter:Devuelve la letra de la columna a la que pertenece la celda.
        # - openpyxl.Cell.coordinate:   Retorna la coordenada de la celda en formato de columna letra + número de 
        #   fila (por ejemplo, "A1").
        # - openpyxl.Cell.data_type:    Retorna el tipo de dato de la celda (por ejemplo, 'n' para numérico, 's' 
        #   para cadena, etc.).
        # - openpyxl.Cell.style:        Retorna el estilo aplicado a la celda.
        # - openpyxl.Cell.hyperlink:    Devuelve la URL de hipervínculo asociada a la celda, si la hay.
        # - openpyxl.Cell.comment:      Devuelve el comentario asociado a la celda, si lo hay.
        # - openpyxl.Cell.font:         Retorna la configuración de fuente (estilo, tamaño, color, etc.) aplicada 
        #   a la celda.
        # - openpyxl.Cell.alignment:    Retorna la configuración de alineación (horizontal y vertical) aplicada a 
        #   la celda.
        # - openpyxl.Cell.border:       Devuelve la configuración de borde aplicada a la celda.
        # - openpyxl.Cell.fill:         Devuelve la configuración de relleno (color, patrón, etc.) aplicada a la 
        #   celda.
        # - openpyxl.Cell.number_format: El formato numérico aplicado a la celda.
        column1Index = cellObject.column
        #openpyxl.load_workbook().active[index]: Con la sintaxis de corchetes aplicada al workSheet de un Excel 
        #podemos acceder a todas las celdas pertenecientes a la columna indicada. Su índice se indica contando 
        #desde 1 y el método devuelve una tupla () que contiene varios elementos que representan cada celda de 
        #forma individual para esa columna, indicando su objeto openpyxl.Cell, la Sheet a la que pertenece y la 
        #coordenada de su celda.
        #TIPOS DE ESTRUCTURAS DE DATOS EN PYTHON: La gran diferencia que estos pueden tener es que algunos tienen 
        #cierto órden (índice y valor) y otros no, además de que algunos son editables o mutables, donde se les 
        #puede agregar, eliminar, o modificar elementos y otros son inmutables, donde sus datos no se pueden 
        #cambiar.
        # - Listas (list): Una lista es una colección ordenada y mutable (editable) de elementos. Se definen 
        #   utilizando corchetes [].
        #       Ejemplo: mi_lista = [1, 2, "hola", True].
        # - Tuplas (tuple): Una tupla es una colección ordenada e inmutable de elementos. Se definen utilizando 
        #   paréntesis ().
        #       Ejemplo: mi_tupla = (1, 2, "hola", True).
        # - Diccionarios (dict): Un diccionario es una colección desordenada y mutable de pares clave-valor. 
        #   Se definen utilizando llaves {} y separando cada par clave-valor por dos puntos :.
        #       Ejemplo: mi_diccionario = {"nombre": "Juan", "edad": 30, "ciudad": "Madrid"}.
        # - Conjuntos (set): Un conjunto es una colección desordenada y mutable de elementos únicos. No permite 
        #   elementos duplicados y no tiene un orden definido. Se definen utilizando llaves {} o usando la 
        #   función set().
        #       Ejemplo: mi_conjunto = {1, 2, 3, 4, 5}.
        col_1_Excel = excelWorkSheet[column1Index]
        print("Celdas de la columna 1:\n", col_1_Excel)
        #Lo que se hará con la variable numLetrasCol_1 es comparar el número de letras que se tenga en todas las 
        #celdas pertenecientes a la columna 1, se empieza con un valor inicial de 0, pero en el bucle posterior 
        #este valor se sobreescribirá y se comparará en cada iteración, para que si el número de letras es mayor 
        #a la iteración anterior, este se ajuste para que así tome el valor mayor del número de letras de la 
        #primera columna.
        numLetrasCol_1 = 0  #Valor inicial del número de letras de las celdas de la primera columna = 0.
        #Bucle for each: Recorre uno a uno todos los elementos de una estructura de datos.
        for cell in (col_1_Excel):
            #Cada elemento de la tupla col_1_Excel es un objeto openpyxl.Cell, por lo cual podemos acceder a 
            #su contenido a través de los parámetros de la clase.
            # - openpyxl.Cell.value: Atributo que devuelve el valor contenido en la celda en forma de string.
            #Si el número de letras de la celda actual es mayor al de la iteración anterior, se ejecuta el 
            #condicional, actualizando el valor de la variable numLetrasCol_1. Además, debemos asegurarnos que el 
            #valor de la celda no está vacío, por eso esa condición también está incluida en el condicional.
            if ((cell.value is not None) and (len(cell.value) > numLetrasCol_1)):
                #Si el contenido string de la celda tiene un número de letras mayor a 0 inicialmente o en 
                #posteriores iteraciones tiene un mayor número de letras que la iteración anterior, se actualiza 
                #el valor de la variable numLetrasCol_1.
                numLetrasCol_1 = len(cell.value)    #Actualización del número máximo de letras.
        #Una vez teniendo el número máximo de letras, se aplica una fórmula matemática que agregará cierto 
        #espacio extra a la celda para que su contenido no se encuentre apachurrado en su celda. Este se puede 
        #modificar al gusto para agregar el espacio extra que se quiera. La suma aumenta el número de letras y la
        #multiplicación ajusta el ancho de la celda al gusto.
        anchoMaxCol1 = (numLetrasCol_1 + 2) * 1.4                       #Ancho en función del número de letras.
        #min(num1, num2, num_n): Método que retorna el valor mínimo al comparar varios números. De esta forma se 
        #puede limitar el ancho de las celdas, comparando el ancho máximo en función del número de letras mayor, 
        #contra el ancho máximo de celda definido por el usuario.
        adjustedWidth_Col1 = min(anchoMaxCol1, self.ancho_maximo)       #Limitación del ancho en las columnas.
        #openpyxl.load_workbook().active.column_dimensions[].width: Este método se utiliza para definir el ancho 
        #de todas las celdas pertecientes a una columna, la cual se indica a través de su letra asignada en el 
        #archivo (workbook) de Excel. El atributo column_dimensions[] pertenece al objeto Worksheet. 
        excelWorkSheet.column_dimensions['A'].width = adjustedWidth_Col1

        #openpyxl.Worksheet: Los atributos del objeto Worksheet que devuelven propiedades específicas son:
        # - openpyxl.Worksheet.title:               Atributo que devuelve el nombre de la hoja de cálculo.
        # - openpyxl.Worksheet.max_row:             Devuelve el número de la fila más alta que contenga datos.
        # - openpyxl.Worksheet.max_column:          Devuelve el número de la columna más a la derecha que 
        #   contenga datos.
        # - openpyxl.Worksheet.dimensions:          Retorna una tupla que contiene el rango de celdas ocupadas 
        #   en la hoja de cálculo (ej. "A1:Z10").
        # - openpyxl.Worksheet.rows:                Retorna un objeto generador que produce todas las filas de 
        #   la hoja de cálculo, incluyendo las celdas vacías. Los generadores son estructuras de datos que 
        #   instancían la clase Generator, estos se pueden recorrer con un bucle for como todos los demás, pero 
        #   en cuestión de memoria son mejores porque no almacenan sus datos de jalón como todos los demás, sino 
        #   que los van creando en la marcha, mientras los vayamos necesitando o accediendo. Si quiero ver sus 
        #   valores, este se deberá convertir a una lista.
        # - openpyxl.Worksheet.columns:             Retorna un generador que produce todas las columnas de la 
        #   hoja de cálculo, incluyendo las celdas vacías.  Si quiero ver sus valores, este se deberá convertir 
        #   a una lista.
        # - openpyxl.Worksheet.cell(row, column):   Método para acceder a una celda específica por su fila y 
        #   columna.
        # - openpyxl.Worksheet.iter_rows():         Método para iterar sobre todas las filas de la hoja de 
        #   cálculo.
        # - openpyxl.Worksheet.iter_cols():         Método para iterar sobre todas las columnas de la hoja de 
        #   cálculo.
        # - openpyxl.Worksheet.column_dimensions:   Devuelve un diccionario que contiene información sobre las 
        #   dimensiones de cada columna, como el ancho de la columna.
        # - openpyxl.Worksheet.row_dimensions:      Devuelve un diccionario que contiene información sobre las 
        #   dimensiones de cada fila, como la altura de la fila.
        # - openpyxl.Worksheet.merge_cells:         Devuelve una propiedad que contiene un conjunto de rangos de 
        #   celdas fusionadas en la hoja de cálculo.
        # - openpyxl.Worksheet.unmerge_cells():     Método para deshacer la fusión de celdas.
        # - openpyxl.Worksheet.freeze_panes:        Propiedad para fijar paneles en una hoja de cálculo.
        # - openpyxl.Worksheet.print_area:          Propiedad que especifica el área de impresión de la hoja de 
        #   cálculo.
        print("Tipo de dato Generator de las columnas:\n", type(excelWorkSheet.columns))    #Objeto Generator.
        #list(): Aunque el objeto Generator es mejor para el manejo de memoria, este no se puede visualizar, por 
        #lo que se debe convertir a una lista para poder manejarlo y visualizarlo. Además, como ya se ajustó el 
        #tamaño de las celdas de la primera columna, debemos tomar el rango desde la columna 2 hasta la final a 
        #través de la sintaxis básica de slicing, los indices se cuentan desde 0, pero el intervalo se cuenta 
        #desde 1:
        # - lista[index_inicio : index_final : intervalo]
        colsExcel = list(excelWorkSheet.columns)[1:]    #Se toman todas las columnas desde la 2 hasta la final.
        #Al convertir el objeto Generator en una lista, podré visualizar una lista de tuplas que contienen 
        #objetos openpyxl.Cell, por lo cual podré acceder a sus atributos, estos me servirán para recorrer todas 
        #las columnas de la tabla.
        print(colsExcel)
        for col in colsExcel:
            numLetrasCols = 0   #Número inicial de las letras en las celdas de las columnas menos la primera = 0.
            #Como colsExcel es una lista de tuplas que contienen objetos openpyxl.Cell, al acceder a su primera 
            #posición [0], accederé a sus tuplas y como me encuentro en un bucle for each, podré acceder 
            #individualmente a los atributos de los objetos Cell que se encuentran dentro.
            #openpyxl.Cell: Los atributos del objeto Cell que devuelven propiedades específicas son:
            # - openpyxl.Cell.column_letter: Devuelve la letra de la columna a la que pertenece la celda.
            column = col[0].column_letter   #Letra de la coordenada de cada columna de la tabla.
            #Este bucle for each compara el número de letras que se tenga en todas las celdas pertenecientes a 
            #las columnas de la tabla, empezando con un valor inicial de 0 letras, pero en cada iteración este 
            #valor se sobreescribirá y se comparará; si el número de letras actual es mayor a la iteración 
            #anterior, se actualizará el valor de la variable numLetrasCols. Además checa si la celda está vacía,
            #si este es el caso, no actualiza el valor de la variable numLetrasCols.
            for cell in col:
                if ((cell.value is not None) and (len(cell.value) > numLetrasCols)):
                    numLetrasCols = len(cell.value)
            #Una vez teniendo el número máximo de letras, se aplica una fórmula matemática que agregará cierto 
            #espacio extra a la celda para que su contenido no se encuentre apachurrado en su celda. Este se 
            #puede modificar al gusto para agregar el espacio extra que se quiera. La suma aumenta el número de 
            #letras y la multiplicación ajusta el ancho de la celda al gusto.
            anchoMaxCols = (numLetrasCols + 2) * 1.3                   #Ancho en función del número de letras.
            #min(num1, num2, num_n): Método que retorna el valor mínimo al comparar varios números. De esta forma 
            #se puede limitar el ancho de las celdas, comparando el ancho máximo en función del número de letras 
            #mayor, contra el ancho máximo de celda definido por el usuario.
            adjusted_width_Cols = min(anchoMaxCols, self.ancho_maximo)  #Limitación del ancho en las columnas.
            #openpyxl.load_workbook().active.column_dimensions[].width: Este método se utiliza para definir el 
            #ancho de todas las celdas pertecientes a una columna, la cual se indica a través de su letra 
            #asignada en el archivo (workbook) de Excel. Debido a esto es que se obtuvo el atributo 
            #openpyxl.Cell.column_letter anteriormente dentro de este mismo bucle for.
            excelWorkSheet.column_dimensions[column].width = adjusted_width_Cols

        #openpyxl.load_workbook().save(path): El método Workbook.save() se utiliza para guardar un archivo de 
        #Excel, especificando su directorio por medio de un path.
        excelWorkBook.save(self.ruta_excel)